#Μεταφορά δεδομένων από PDF σε excel.
#Ονομάζουμε το τιμολόγιό μας σε 'invoice.pdf' και το βάζουμε στο φάκελο του py
#Βάζουμε RegEx στις παρενθέσεις για να αναγνωρίζει όλα τα τιμολόγια του προμηθευτή

from PyPDF2 import PdfReader
import re
import openpyxl

# Open the PDF invoice using PyPDF2
pdf = PdfReader('invoice.pdf')

# Extract text from PDF
text = ''
for page in pdf.pages:
    text += page.extract_text()

# Use regular expression or string matching to find invoice number and amount
invoice_number = re.findall(r'INVOICE NO. (G2M-23-24-E-060)', text)[0]
invoice_date = re.findall(r'INVOICE DATE (26‐Dec‐2023)', text)[0]
total_amount = re.findall(r'Total EX‐WORKS Price : (€ 1,500.00)', text)[0]

# Write the extracted values to Excel using openpyxl
excel_file = openpyxl.load_workbook('invoices.xlsx')
sheet = excel_file.active
row = sheet.max_row + 1
sheet.cell(row=row, column=1).value = invoice_number
sheet.cell(row=row, column=1).value = invoice_date
sheet.cell(row=row, column=2).value = total_amount

excel_file.save('invoices.xlsx')